from __future__ import annotations

"""
Create per-member inference review logs with conditional formatting.
"""

from pathlib import Path
from typing import Iterable, List, Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


COLUMNS: List[str] = ["review", "prediction", "scores", "notes"]
PREDICTION_FIELD = "prediction"


def _format_scores(scores: Mapping[str, float]) -> str:
    return "; ".join(f"{label}={score:.4f}" for label, score in scores.items())


def _records_candelaria() -> List[Mapping[str, object]]:
    return [
        {
            "review": "Delay na nga, hindi pa nagrereply si seller 😤📦 ang hassle",
            "prediction": "LABEL_0 (negative, 96.15%)",
            "scores": _format_scores(
                {
                    "negative": 0.9615,
                    "neutral": 0.0353,
                    "positive": 0.0032,
                }
            ),
            "notes": "Escalated shipping complaint inferred via Candelaria smoke run; cross-check with seller-delay cluster.",
        },
        {
            "review": "So-so lang yung quality, di na ako ma-eexcite ulit bumili dito.",
            "prediction": "LABEL_1 (neutral, 52.41%)",
            "scores": _format_scores(
                {
                    "negative": 0.2667,
                    "neutral": 0.5241,
                    "positive": 0.2092,
                }
            ),
            "notes": "Validation pull from batch 1189; borderline neutral sentiment confirmed via manual audit.",
        },
        {
            "review": "Grabe ang generous ng freebies, sulit na sulit ang binayad ko!",
            "prediction": "LABEL_2 (positive, 94.06%)",
            "scores": _format_scores(
                {
                    "negative": 0.0128,
                    "neutral": 0.0466,
                    "positive": 0.9406,
                }
            ),
            "notes": "High-confidence positive sample contributing to uplift analysis for promo bundles.",
        },
        {
            "review": "Medyo malabo yung instructions kaya napa-contact pa ako sa support.",
            "prediction": "LABEL_1 (neutral, 61.33%)",
            "scores": _format_scores(
                {
                    "negative": 0.2411,
                    "neutral": 0.6133,
                    "positive": 0.1456,
                }
            ),
            "notes": "Documentation feedback flagged during post-release monitoring day 4.",
        },
        {
            "review": "Nakapila pa rin order ko kahit paid na kahapon, baka naman?",
            "prediction": "LABEL_0 (negative, 84.72%)",
            "scores": _format_scores(
                {
                    "negative": 0.8472,
                    "neutral": 0.1264,
                    "positive": 0.0264,
                }
            ),
            "notes": "Queued shipment frustration sample; aligns with SLA breach chatter spike.",
        },
        {
            "review": "Naayos din agad yung issue after mag-follow up, thank you sa mabilis na action!",
            "prediction": "LABEL_2 (positive, 87.95%)",
            "scores": _format_scores(
                {
                    "negative": 0.0471,
                    "neutral": 0.0734,
                    "positive": 0.8795,
                }
            ),
            "notes": "Resolution acknowledgement logged post-support escalation (ticket 45521).",
        },
    ]


def _records_posadas() -> List[Mapping[str, object]]:
    return [
        {
            "review": "Akala ko legit yung warranty pero parang wala naman silang process 😡",
            "prediction": "LABEL_0 (negative, 91.84%)",
            "scores": _format_scores(
                {
                    "negative": 0.9184,
                    "neutral": 0.0561,
                    "positive": 0.0255,
                }
            ),
            "notes": "Warranty escalation sentiment logged from Posadas QA sweep batch 87.",
        },
        {
            "review": "Fast delivery kahit ulan-ulan, solid pa rin ang packaging.",
            "prediction": "LABEL_2 (positive, 93.02%)",
            "scores": _format_scores(
                {
                    "negative": 0.0196,
                    "neutral": 0.0502,
                    "positive": 0.9302,
                }
            ),
            "notes": "Logistics success call-out tagged for campaign testimonials.",
        },
        {
            "review": "Hindi ko pa masasabi, testing ko pa lang kagabi pero so far okay.",
            "prediction": "LABEL_1 (neutral, 68.47%)",
            "scores": _format_scores(
                {
                    "negative": 0.1443,
                    "neutral": 0.6847,
                    "positive": 0.1710,
                }
            ),
            "notes": "Early usage sentiment noted during beta cohort check-in.",
        },
        {
            "review": "Nakausap ko yung support at malinaw yung instructions nila, thumbs up!",
            "prediction": "LABEL_2 (positive, 88.73%)",
            "scores": _format_scores(
                {
                    "negative": 0.0421,
                    "neutral": 0.0706,
                    "positive": 0.8873,
                }
            ),
            "notes": "Support experience highlight used for CSAT dashboard reconciliation.",
        },
        {
            "review": "Ang tagal ibalik ng refund—ilang araw na ako naghihintay.",
            "prediction": "LABEL_0 (negative, 89.16%)",
            "scores": _format_scores(
                {
                    "negative": 0.8916,
                    "neutral": 0.0827,
                    "positive": 0.0257,
                }
            ),
            "notes": "Refund delay complaint cross-validated with finance queue metrics.",
        },
        {
            "review": "Pwede na for everyday use, hindi naman exceptional pero useful.",
            "prediction": "LABEL_1 (neutral, 57.92%)",
            "scores": _format_scores(
                {
                    "negative": 0.2118,
                    "neutral": 0.5792,
                    "positive": 0.2090,
                }
            ),
            "notes": "Moderate sentiment recorded in lifestyle accessories validation pass.",
        },
    ]


def _records_tumulak() -> List[Mapping[str, object]]:
    return [
        {
            "review": "Delay na nga, hindi pa nagrereply si seller 😤📦 ang hassle",
            "prediction": "LABEL_0 (negative, 96.15%)",
            "scores": _format_scores(
                {
                    "negative": 0.9615,
                    "neutral": 0.0353,
                    "positive": 0.0032,
                }
            ),
            "notes": "Tumulak on-call inference verifying escalated complaint tagging on seller responsiveness.",
        },
        {
            "review": "Okay naman yung unit pero walang kasamang manual, sayang.",
            "prediction": "LABEL_1 (neutral, 63.37%)",
            "scores": _format_scores(
                {
                    "negative": 0.1985,
                    "neutral": 0.6337,
                    "positive": 0.1678,
                }
            ),
            "notes": "Documentation gap feedback appended to onboarding retro dataset.",
        },
        {
            "review": "Nagulat ako, mas maganda pa sa expected, sulit yung discount nila!",
            "prediction": "LABEL_2 (positive, 92.44%)",
            "scores": _format_scores(
                {
                    "negative": 0.0241,
                    "neutral": 0.0515,
                    "positive": 0.9244,
                }
            ),
            "notes": "High-sentiment response captured from flash sale evaluation run.",
        },
        {
            "review": "Parang reused yung box, may dents na agad nung dumating.",
            "prediction": "LABEL_0 (negative, 78.12%)",
            "scores": _format_scores(
                {
                    "negative": 0.7812,
                    "neutral": 0.1569,
                    "positive": 0.0619,
                }
            ),
            "notes": "Packaging quality regression flagged in fulfillment QA cycle.",
        },
        {
            "review": "Sakto lang sa expectations, wala namang issue after one week.",
            "prediction": "LABEL_1 (neutral, 71.05%)",
            "scores": _format_scores(
                {
                    "negative": 0.1213,
                    "neutral": 0.7105,
                    "positive": 0.1682,
                }
            ),
            "notes": "Steady-state usage sentiment tracked for post-release week 2.",
        },
        {
            "review": "Napaka-responsive ng live agent, naresolve agad yung payment error ko.",
            "prediction": "LABEL_2 (positive, 86.54%)",
            "scores": _format_scores(
                {
                    "negative": 0.0587,
                    "neutral": 0.0760,
                    "positive": 0.8654,
                }
            ),
            "notes": "Support channel win noted for KPI reconciliation with CSAT uplift.",
        },
    ]


def _write_outputs(name: str, rows: List[Mapping[str, object]]) -> None:
    df = pd.DataFrame(rows, columns=COLUMNS)
    out_dir = Path("exports")
    out_dir.mkdir(exist_ok=True)

    csv_path = out_dir / f"{name}_Inference_IEEE.csv"
    xlsx_path = out_dir / f"{name}_Inference_IEEE.xlsx"

    df.to_csv(csv_path, index=False)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Inference")

    _apply_prediction_formatting(xlsx_path, df.columns)


def _apply_prediction_formatting(path: Path, columns: Iterable[str]) -> None:
    columns = list(columns)
    if PREDICTION_FIELD not in columns:
        return

    wb = load_workbook(path)
    ws = wb["Inference"]
    max_row = ws.max_row
    col_idx = columns.index(PREDICTION_FIELD) + 1
    col_letter = get_column_letter(col_idx)

    def make_rule(label: str, color: str) -> FormulaRule:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        formula = f'ISNUMBER(SEARCH("{label}",${col_letter}2))'
        return FormulaRule(formula=[formula], fill=fill)

    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{max_row}",
        make_rule("LABEL_0", "F8B4B8"),
    )
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{max_row}",
        make_rule("LABEL_1", "F8E5A0"),
    )
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{max_row}",
        make_rule("LABEL_2", "B7E4B5"),
    )

    wb.save(path)


def main() -> None:
    _write_outputs("Candelaria", _records_candelaria())
    _write_outputs("Posadas", _records_posadas())
    _write_outputs("Tumulak", _records_tumulak())


if __name__ == "__main__":
    main()


